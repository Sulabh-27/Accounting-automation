"""
MIS Generator Agent for Part-8: MIS & Audit Trail

Generates Management Information System reports by aggregating sales, expenses,
and computing business intelligence metrics from processed e-commerce data.
"""

import os
import uuid
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional, List
from dataclasses import dataclass

from ..libs.supabase_client import SupabaseClientWrapper
from ..libs.mis_utils import MISCalculator, MISReport
from ..libs.audit_utils import AuditLogger, AuditActor, AuditAction
from ..libs.utils import ensure_dir


@dataclass
class MISGenerationResult:
    """Result of MIS generation process"""
    success: bool
    mis_report: Optional[MISReport]
    report_id: Optional[str]
    csv_export_path: Optional[str]
    excel_export_path: Optional[str]
    error_message: Optional[str]
    processing_time_seconds: float


class MISGeneratorAgent:
    """
    Management Information System Generator Agent
    
    Features:
    - Aggregates sales data from pivot summaries
    - Processes expense data from seller invoices
    - Computes GST liability and tax metrics
    - Calculates profitability and margin analysis
    - Generates comprehensive business intelligence reports
    - Exports reports in multiple formats (CSV, Excel, Database)
    - Integrates with audit logging for traceability
    """
    
    def __init__(self, supabase_client: Optional[SupabaseClientWrapper] = None):
        self.supabase = supabase_client or SupabaseClientWrapper()
        self.mis_calculator = MISCalculator(self.supabase)
        self.audit_logger = AuditLogger(self.supabase)
        
        # Create MIS export directories
        self.mis_export_dir = "ingestion_layer/exports/mis"
        ensure_dir(self.mis_export_dir)
    
    def generate_mis_report(
        self,
        run_id: uuid.UUID,
        channel: str,
        gstin: str,
        month: str,
        report_type: str = "monthly",
        export_formats: List[str] = ["csv", "database"]
    ) -> MISGenerationResult:
        """
        Generate comprehensive MIS report
        
        Args:
            run_id: Processing run identifier
            channel: E-commerce channel (amazon, flipkart, pepperfry)
            gstin: Company GSTIN
            month: Processing month (YYYY-MM format)
            report_type: Report type (monthly, quarterly, annual)
            export_formats: List of export formats (csv, excel, database)
            
        Returns:
            MIS generation result with report data and export paths
        """
        start_time = datetime.now()
        
        try:
            print(f"📊 Starting MIS report generation...")
            print(f"   🏢 Channel: {channel}")
            print(f"   🏛️  GSTIN: {gstin}")
            print(f"   📅 Month: {month}")
            print(f"   📋 Report Type: {report_type}")
            
            # Log MIS generation start
            self.audit_logger.log_event(
                run_id=run_id,
                actor=AuditActor.SYSTEM,
                action=AuditAction.MIS_GENERATED,
                details={
                    'stage': 'start',
                    'channel': channel,
                    'gstin': gstin,
                    'month': month,
                    'report_type': report_type
                }
            )
            
            # Generate MIS report using calculator
            mis_report = self.mis_calculator.generate_mis_report(
                run_id=run_id,
                channel=channel,
                gstin=gstin,
                month=month,
                report_type=report_type
            )
            
            # Export report in requested formats
            csv_export_path = None
            excel_export_path = None
            report_id = None
            
            # Export to database
            if "database" in export_formats:
                report_id = self.mis_calculator.save_mis_report(mis_report)
                print(f"💾 MIS report saved to database: {report_id}")
            
            # Export to CSV
            if "csv" in export_formats:
                csv_filename = f"mis_report_{channel}_{gstin}_{month}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                csv_export_path = os.path.join(self.mis_export_dir, csv_filename)
                self.mis_calculator.export_mis_report_csv(mis_report, csv_export_path)
                print(f"📄 MIS report exported to CSV: {csv_export_path}")
            
            # Export to Excel
            if "excel" in export_formats:
                excel_filename = f"mis_report_{channel}_{gstin}_{month}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_export_path = os.path.join(self.mis_export_dir, excel_filename)
                self._export_mis_report_excel(mis_report, excel_export_path)
                print(f"📊 MIS report exported to Excel: {excel_export_path}")
            
            # Calculate processing time
            end_time = datetime.now()
            processing_time = (end_time - start_time).total_seconds()
            
            # Log MIS generation completion
            self.audit_logger.log_event(
                run_id=run_id,
                actor=AuditActor.SYSTEM,
                action=AuditAction.MIS_GENERATED,
                details={
                    'stage': 'complete',
                    'channel': channel,
                    'gstin': gstin,
                    'month': month,
                    'report_type': report_type,
                    'report_id': report_id,
                    'csv_export_path': csv_export_path,
                    'excel_export_path': excel_export_path,
                    'processing_time_seconds': processing_time,
                    'metrics': {
                        'total_sales': float(mis_report.sales_metrics.total_sales),
                        'total_expenses': float(mis_report.expense_metrics.total_expenses),
                        'gross_profit': float(mis_report.profitability_metrics.gross_profit),
                        'profit_margin': float(mis_report.profitability_metrics.profit_margin),
                        'gst_liability': float(mis_report.gst_metrics.gst_liability),
                        'data_quality_score': float(mis_report.data_quality_score)
                    }
                }
            )
            
            # Flush audit logs
            self.audit_logger.flush_logs()
            
            print(f"✅ MIS report generation completed successfully")
            print(f"   ⏱️  Processing time: {processing_time:.2f} seconds")
            print(f"   📈 Key metrics:")
            print(f"      💰 Net Sales: ₹{mis_report.sales_metrics.net_sales:,.2f}")
            print(f"      💸 Total Expenses: ₹{mis_report.expense_metrics.total_expenses:,.2f}")
            print(f"      📊 Gross Profit: ₹{mis_report.profitability_metrics.gross_profit:,.2f}")
            print(f"      📈 Profit Margin: {mis_report.profitability_metrics.profit_margin:.1f}%")
            print(f"      🏛️  GST Liability: ₹{mis_report.gst_metrics.gst_liability:,.2f}")
            print(f"      ⭐ Quality Score: {mis_report.data_quality_score:.1f}%")
            
            return MISGenerationResult(
                success=True,
                mis_report=mis_report,
                report_id=report_id,
                csv_export_path=csv_export_path,
                excel_export_path=excel_export_path,
                error_message=None,
                processing_time_seconds=processing_time
            )
            
        except Exception as e:
            error_message = f"MIS generation failed: {str(e)}"
            processing_time = (datetime.now() - start_time).total_seconds()
            
            print(f"❌ {error_message}")
            
            # Log MIS generation error
            self.audit_logger.log_event(
                run_id=run_id,
                actor=AuditActor.SYSTEM,
                action=AuditAction.CRITICAL_ERROR,
                details={
                    'stage': 'error',
                    'channel': channel,
                    'gstin': gstin,
                    'month': month,
                    'error_message': error_message,
                    'processing_time_seconds': processing_time
                }
            )
            
            self.audit_logger.flush_logs()
            
            return MISGenerationResult(
                success=False,
                mis_report=None,
                report_id=None,
                csv_export_path=None,
                excel_export_path=None,
                error_message=error_message,
                processing_time_seconds=processing_time
            )
    
    def generate_comparative_report(
        self,
        channel: str,
        gstin: str,
        months: List[str],
        report_type: str = "comparative"
    ) -> Dict[str, Any]:
        """
        Generate comparative MIS report across multiple months
        
        Args:
            channel: E-commerce channel
            gstin: Company GSTIN
            months: List of months to compare (YYYY-MM format)
            report_type: Report type
            
        Returns:
            Comparative report data
        """
        try:
            print(f"📊 Generating comparative MIS report...")
            print(f"   🏢 Channel: {channel}")
            print(f"   🏛️  GSTIN: {gstin}")
            print(f"   📅 Months: {', '.join(months)}")
            
            comparative_data = []
            
            # Get MIS reports for each month
            for month in months:
                try:
                    if hasattr(self.supabase, 'client') and self.supabase.client:
                        result = self.supabase.client.table('mis_reports').select('*').eq('channel', channel).eq('gstin', gstin).eq('month', month).execute()
                        if result.data:
                            comparative_data.append(result.data[0])
                        else:
                            print(f"⚠️  No MIS data found for {month}")
                    else:
                        print(f"⚠️  Development mode - no comparative data available for {month}")
                        
                except Exception as e:
                    print(f"⚠️  Error retrieving data for {month}: {e}")
            
            if not comparative_data:
                print("❌ No comparative data available")
                return {'error': 'No comparative data available'}
            
            # Create comparative analysis
            comparison = {
                'channel': channel,
                'gstin': gstin,
                'months': months,
                'report_type': report_type,
                'data': comparative_data,
                'trends': self._calculate_trends(comparative_data),
                'generated_at': datetime.now().isoformat()
            }
            
            # Export comparative report
            comparative_filename = f"comparative_report_{channel}_{gstin}_{'-'.join(months)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            comparative_path = os.path.join(self.mis_export_dir, comparative_filename)
            self._export_comparative_report_csv(comparison, comparative_path)
            
            print(f"✅ Comparative report generated: {comparative_path}")
            
            return comparison
            
        except Exception as e:
            error_message = f"Comparative report generation failed: {str(e)}"
            print(f"❌ {error_message}")
            return {'error': error_message}
    
    def get_mis_dashboard_data(
        self,
        channel: Optional[str] = None,
        gstin: Optional[str] = None,
        limit: int = 10
    ) -> Dict[str, Any]:
        """
        Get MIS dashboard data for business intelligence
        
        Args:
            channel: Optional channel filter
            gstin: Optional GSTIN filter
            limit: Number of recent reports to include
            
        Returns:
            Dashboard data with key metrics and trends
        """
        try:
            print(f"📊 Retrieving MIS dashboard data...")
            
            if hasattr(self.supabase, 'client') and self.supabase.client:
                query = self.supabase.client.table('mis_reports').select('*')
                
                if channel:
                    query = query.eq('channel', channel)
                if gstin:
                    query = query.eq('gstin', gstin)
                
                result = query.order('created_at', desc=True).limit(limit).execute()
                reports = result.data
            else:
                print("⚠️  Development mode - no dashboard data available")
                reports = []
            
            if not reports:
                return {'error': 'No MIS reports available'}
            
            # Calculate dashboard metrics
            dashboard_data = {
                'summary': {
                    'total_reports': len(reports),
                    'channels': list(set(r['channel'] for r in reports)),
                    'gstins': list(set(r['gstin'] for r in reports)),
                    'latest_month': reports[0]['month'] if reports else None
                },
                'recent_reports': reports,
                'aggregated_metrics': self._calculate_aggregated_metrics(reports),
                'trends': self._calculate_dashboard_trends(reports),
                'generated_at': datetime.now().isoformat()
            }
            
            print(f"✅ Dashboard data retrieved: {len(reports)} reports")
            
            return dashboard_data
            
        except Exception as e:
            error_message = f"Dashboard data retrieval failed: {str(e)}"
            print(f"❌ {error_message}")
            return {'error': error_message}
    
    def _export_mis_report_excel(self, mis_report: MISReport, output_path: str) -> str:
        """Export MIS report to Excel format with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add summary data
            summary_data = [
                ["MIS Report Summary", ""],
                ["Channel", mis_report.channel],
                ["GSTIN", mis_report.gstin],
                ["Month", mis_report.month],
                ["Report Type", mis_report.report_type],
                ["Generated At", mis_report.created_at.strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["Sales Metrics", ""],
                ["Total Sales", f"₹{mis_report.sales_metrics.total_sales:,.2f}"],
                ["Total Returns", f"₹{mis_report.sales_metrics.total_returns:,.2f}"],
                ["Net Sales", f"₹{mis_report.sales_metrics.net_sales:,.2f}"],
                ["Total Transactions", mis_report.sales_metrics.total_transactions],
                ["Total SKUs", mis_report.sales_metrics.total_skus],
                ["", ""],
                ["Expense Metrics", ""],
                ["Total Expenses", f"₹{mis_report.expense_metrics.total_expenses:,.2f}"],
                ["Commission Expenses", f"₹{mis_report.expense_metrics.commission_expenses:,.2f}"],
                ["Shipping Expenses", f"₹{mis_report.expense_metrics.shipping_expenses:,.2f}"],
                ["Fulfillment Expenses", f"₹{mis_report.expense_metrics.fulfillment_expenses:,.2f}"],
                ["", ""],
                ["Profitability Metrics", ""],
                ["Gross Profit", f"₹{mis_report.profitability_metrics.gross_profit:,.2f}"],
                ["Profit Margin", f"{mis_report.profitability_metrics.profit_margin:.2f}%"],
                ["", ""],
                ["GST Metrics", ""],
                ["GST Output", f"₹{mis_report.gst_metrics.net_gst_output:,.2f}"],
                ["GST Input", f"₹{mis_report.gst_metrics.net_gst_input:,.2f}"],
                ["GST Liability", f"₹{mis_report.gst_metrics.gst_liability:,.2f}"],
                ["", ""],
                ["Quality Metrics", ""],
                ["Data Quality Score", f"{mis_report.data_quality_score:.1f}%"],
                ["Exception Count", mis_report.exception_count],
                ["Approval Count", mis_report.approval_count]
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)
                
                # Style headers
                if label and not value:
                    cell = ws_summary.cell(row=row_idx, column=1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_path)
            return output_path
            
        except ImportError:
            print("⚠️  openpyxl not available - skipping Excel export")
            return output_path
        except Exception as e:
            print(f"⚠️  Error exporting to Excel: {e}")
            return output_path
    
    def _export_comparative_report_csv(self, comparison: Dict[str, Any], output_path: str) -> str:
        """Export comparative report to CSV"""
        try:
            # Create DataFrame from comparative data
            df = pd.DataFrame(comparison['data'])
            
            # Select key columns for comparison
            comparison_columns = [
                'month', 'total_sales', 'total_expenses', 'gross_profit', 
                'profit_margin', 'net_gst_output', 'net_gst_input', 'gst_liability',
                'total_transactions', 'data_quality_score'
            ]
            
            # Filter to available columns
            available_columns = [col for col in comparison_columns if col in df.columns]
            df_export = df[available_columns]
            
            # Export to CSV
            df_export.to_csv(output_path, index=False)
            return output_path
            
        except Exception as e:
            print(f"⚠️  Error exporting comparative report: {e}")
            return output_path
    
    def _calculate_trends(self, comparative_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate trends from comparative data"""
        try:
            if len(comparative_data) < 2:
                return {'error': 'Insufficient data for trend analysis'}
            
            # Sort by month
            sorted_data = sorted(comparative_data, key=lambda x: x['month'])
            
            # Calculate trends
            first_month = sorted_data[0]
            last_month = sorted_data[-1]
            
            trends = {
                'sales_growth': self._calculate_growth_rate(first_month['total_sales'], last_month['total_sales']),
                'expense_growth': self._calculate_growth_rate(first_month['total_expenses'], last_month['total_expenses']),
                'profit_growth': self._calculate_growth_rate(first_month['gross_profit'], last_month['gross_profit']),
                'transaction_growth': self._calculate_growth_rate(first_month['total_transactions'], last_month['total_transactions'])
            }
            
            return trends
            
        except Exception as e:
            return {'error': f'Trend calculation failed: {str(e)}'}
    
    def _calculate_aggregated_metrics(self, reports: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate aggregated metrics from multiple reports"""
        try:
            if not reports:
                return {}
            
            total_sales = sum(float(r['total_sales']) for r in reports)
            total_expenses = sum(float(r['total_expenses']) for r in reports)
            total_transactions = sum(int(r['total_transactions']) for r in reports)
            
            return {
                'total_sales': total_sales,
                'total_expenses': total_expenses,
                'total_profit': total_sales - total_expenses,
                'total_transactions': total_transactions,
                'average_quality_score': sum(float(r['data_quality_score']) for r in reports) / len(reports),
                'report_count': len(reports)
            }
            
        except Exception as e:
            return {'error': f'Aggregation failed: {str(e)}'}
    
    def _calculate_dashboard_trends(self, reports: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate dashboard trends"""
        try:
            if len(reports) < 2:
                return {'error': 'Insufficient data for trend analysis'}
            
            # Sort by creation date
            sorted_reports = sorted(reports, key=lambda x: x['created_at'])
            
            # Calculate month-over-month trends for recent reports
            recent_trends = {}
            if len(sorted_reports) >= 2:
                prev_report = sorted_reports[-2]
                curr_report = sorted_reports[-1]
                
                recent_trends = {
                    'sales_trend': self._calculate_growth_rate(prev_report['total_sales'], curr_report['total_sales']),
                    'profit_trend': self._calculate_growth_rate(prev_report['gross_profit'], curr_report['gross_profit']),
                    'quality_trend': float(curr_report['data_quality_score']) - float(prev_report['data_quality_score'])
                }
            
            return recent_trends
            
        except Exception as e:
            return {'error': f'Dashboard trend calculation failed: {str(e)}'}
    
    def _calculate_growth_rate(self, old_value: float, new_value: float) -> float:
        """Calculate growth rate percentage"""
        try:
            old_val = float(old_value)
            new_val = float(new_value)
            
            if old_val == 0:
                return 100.0 if new_val > 0 else 0.0
            
            return ((new_val - old_val) / old_val) * 100
            
        except (ValueError, ZeroDivisionError):
            return 0.0
