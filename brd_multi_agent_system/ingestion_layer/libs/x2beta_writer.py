"""
X2Beta Writer Library
Handles reading X2Beta templates and writing batch data to Excel files for Tally import
"""
import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from decimal import Decimal, ROUND_HALF_UP


class X2BetaWriter:
    """
    Utility class for writing batch CSV data to X2Beta Excel templates.
    Handles template loading, field mapping, and Excel formatting.
    """
    
    def __init__(self):
        self.template_cache = {}
        
        # Standard X2Beta column mapping
        self.column_mapping = {
            # Core voucher fields
            'voucher_date': 'Date',
            'voucher_number': 'Voucher No.',
            'voucher_type': 'Voucher Type',
            
            # Party information
            'party_ledger': 'Party Ledger',
            'party_name': 'Party Name',
            'party_gstin': 'Party GSTIN',
            
            # Item details
            'item_name': 'Item Name',
            'item_code': 'Item Code',
            'quantity': 'Quantity',
            'rate': 'Rate',
            'amount': 'Amount',
            
            # Tax ledgers
            'sales_ledger': 'Sales Ledger',
            'cgst_ledger': 'Output CGST Ledger',
            'sgst_ledger': 'Output SGST Ledger',
            'igst_ledger': 'Output IGST Ledger',
            
            # Tax amounts
            'taxable_amount': 'Taxable Amount',
            'cgst_amount': 'CGST Amount',
            'sgst_amount': 'SGST Amount',
            'igst_amount': 'IGST Amount',
            'total_amount': 'Total Amount',
            
            # Additional fields
            'narration': 'Narration',
            'reference': 'Reference'
        }
    
    def load_template(self, template_path: str) -> Workbook:
        """
        Load X2Beta template Excel file.
        
        Args:
            template_path: Path to the X2Beta template file
            
        Returns:
            Loaded workbook object
            
        Raises:
            FileNotFoundError: If template file doesn't exist
        """
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"X2Beta template not found: {template_path}")
        
        # Cache templates to avoid repeated loading
        if template_path not in self.template_cache:
            self.template_cache[template_path] = load_workbook(template_path)
        
        # Return a copy for modification
        return load_workbook(template_path)
    
    def create_default_template(self, gstin: str, company_name: str) -> Workbook:
        """
        Create a default X2Beta template if none exists.
        
        Args:
            gstin: Company GSTIN
            company_name: Company name
            
        Returns:
            New workbook with X2Beta structure
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Vouchers"
        
        # Header row
        headers = [
            'Date', 'Voucher No.', 'Voucher Type', 'Party Ledger', 'Party Name',
            'Item Name', 'Quantity', 'Rate', 'Taxable Amount',
            'Output CGST Ledger', 'CGST Amount',
            'Output SGST Ledger', 'SGST Amount', 
            'Output IGST Ledger', 'IGST Amount',
            'Total Amount', 'Narration'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add company info in first few rows (optional)
        ws.insert_rows(1, 3)
        ws.cell(row=1, column=1, value=f"Company: {company_name}")
        ws.cell(row=2, column=1, value=f"GSTIN: {gstin}")
        ws.cell(row=3, column=1, value="X2Beta Sales Import Template")
        
        return wb
    
    def validate_batch_data(self, df: pd.DataFrame) -> Dict[str, any]:
        """
        Validate batch CSV data before writing to X2Beta template.
        
        Args:
            df: Batch data DataFrame
            
        Returns:
            Validation result with success flag and errors
        """
        errors = []
        
        # Check required columns
        required_columns = ['gst_rate', 'ledger_name', 'fg', 'total_quantity', 'total_taxable']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            errors.append(f"Missing required columns: {missing_columns}")
        
        # Check GST rate consistency
        if 'gst_rate' in df.columns:
            unique_rates = df['gst_rate'].nunique()
            if unique_rates > 1:
                errors.append(f"Multiple GST rates found in batch: {df['gst_rate'].unique()}")
        
        # Check for null values in critical fields
        if 'total_taxable' in df.columns:
            null_taxable = df['total_taxable'].isnull().sum()
            if null_taxable > 0:
                errors.append(f"{null_taxable} records have null taxable amounts")
        
        # Check for negative amounts
        numeric_columns = ['total_quantity', 'total_taxable', 'total_cgst', 'total_sgst', 'total_igst']
        for col in numeric_columns:
            if col in df.columns:
                negative_count = (df[col] < 0).sum()
                if negative_count > 0:
                    errors.append(f"{negative_count} records have negative {col}")
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'record_count': len(df),
            'gst_rate': df['gst_rate'].iloc[0] if 'gst_rate' in df.columns and len(df) > 0 else None
        }
    
    def map_batch_to_x2beta(self, df: pd.DataFrame, gstin: str, month: str) -> pd.DataFrame:
        """
        Map batch CSV data to X2Beta format.
        
        Args:
            df: Batch data DataFrame
            gstin: Company GSTIN
            month: Processing month
            
        Returns:
            DataFrame formatted for X2Beta template
        """
        x2beta_data = []
        
        for i, row in df.iterrows():
            # Generate voucher date (first day of month)
            voucher_date = datetime.strptime(f"{month}-01", "%Y-%m-%d").strftime("%d-%m-%Y")
            
            # Generate voucher number (use existing invoice_no if available)
            voucher_no = row.get('invoice_no', f"SL{month.replace('-', '')}{i+1:04d}")
            
            # Determine tax type
            is_intrastate = row.get('total_cgst', 0) > 0
            is_interstate = row.get('total_igst', 0) > 0
            
            # Map to X2Beta format
            x2beta_record = {
                'Date': voucher_date,
                'Voucher No.': voucher_no,
                'Voucher Type': 'Sales',
                'Party Ledger': row.get('ledger_name', 'Unknown Customer'),
                'Party Name': row.get('party_name', row.get('ledger_name', 'Unknown Customer')),
                'Item Name': row.get('fg', 'Unknown Item'),
                'Quantity': self._round_decimal(row.get('total_quantity', 0)),
                'Rate': self._round_decimal(row.get('total_taxable', 0) / max(row.get('total_quantity', 1), 1)),
                'Taxable Amount': self._round_decimal(row.get('total_taxable', 0)),
                'Total Amount': self._round_decimal(
                    row.get('total_taxable', 0) + 
                    row.get('total_cgst', 0) + 
                    row.get('total_sgst', 0) + 
                    row.get('total_igst', 0)
                ),
                'Narration': f"Sales - {row.get('fg', 'Item')} - {month}"
            }
            
            # Add tax ledgers and amounts based on tax type
            if is_intrastate:
                x2beta_record.update({
                    'Output CGST Ledger': f'Output CGST @ {row.get("gst_rate", 0)*100:.0f}%',
                    'CGST Amount': self._round_decimal(row.get('total_cgst', 0)),
                    'Output SGST Ledger': f'Output SGST @ {row.get("gst_rate", 0)*100:.0f}%',
                    'SGST Amount': self._round_decimal(row.get('total_sgst', 0)),
                    'Output IGST Ledger': '',
                    'IGST Amount': 0
                })
            elif is_interstate:
                x2beta_record.update({
                    'Output CGST Ledger': '',
                    'CGST Amount': 0,
                    'Output SGST Ledger': '',
                    'SGST Amount': 0,
                    'Output IGST Ledger': f'Output IGST @ {row.get("gst_rate", 0)*100:.0f}%',
                    'IGST Amount': self._round_decimal(row.get('total_igst', 0))
                })
            else:
                # Zero GST
                x2beta_record.update({
                    'Output CGST Ledger': '',
                    'CGST Amount': 0,
                    'Output SGST Ledger': '',
                    'SGST Amount': 0,
                    'Output IGST Ledger': '',
                    'IGST Amount': 0
                })
            
            x2beta_data.append(x2beta_record)
        
        return pd.DataFrame(x2beta_data)
    
    def write_to_template(self, 
                         df: pd.DataFrame, 
                         template_path: str, 
                         output_path: str,
                         start_row: int = 5) -> Dict[str, any]:
        """
        Write batch data to X2Beta template and save as new Excel file.
        
        Args:
            df: X2Beta formatted DataFrame
            template_path: Path to X2Beta template
            output_path: Path for output Excel file
            start_row: Row to start writing data (after headers)
            
        Returns:
            Write result with success flag and metadata
        """
        try:
            # Load template
            if os.path.exists(template_path):
                wb = self.load_template(template_path)
                ws = wb.active
            else:
                # Create default template
                wb = self.create_default_template("Unknown", "Unknown Company")
                ws = wb.active
                start_row = 5  # Account for header rows
            
            # Clear existing data (keep headers)
            if ws.max_row > start_row:
                ws.delete_rows(start_row, ws.max_row - start_row + 1)
            
            # Write data starting from start_row
            for r_idx, (_, row) in enumerate(df.iterrows(), start_row):
                for c_idx, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Apply formatting based on data type
                    if isinstance(value, (int, float, Decimal)) and value != 0:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    elif col_name == 'Date':
                        cell.number_format = 'DD-MM-YYYY'
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.alignment = Alignment(horizontal='left')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            
            # Calculate summary statistics
            total_taxable = df['Taxable Amount'].sum() if 'Taxable Amount' in df.columns else 0
            total_tax = (
                df.get('CGST Amount', pd.Series([0])).sum() +
                df.get('SGST Amount', pd.Series([0])).sum() +
                df.get('IGST Amount', pd.Series([0])).sum()
            )
            
            return {
                'success': True,
                'output_path': output_path,
                'record_count': len(df),
                'total_taxable': float(total_taxable),
                'total_tax': float(total_tax),
                'file_size': os.path.getsize(output_path) if os.path.exists(output_path) else 0
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'output_path': output_path,
                'record_count': 0,
                'total_taxable': 0,
                'total_tax': 0,
                'file_size': 0
            }
    
    def _round_decimal(self, value: float, places: int = 2) -> float:
        """
        Round decimal using banker's rounding.
        
        Args:
            value: Value to round
            places: Decimal places
            
        Returns:
            Rounded value
        """
        if pd.isna(value) or value is None:
            return 0.0
        
        decimal_value = Decimal(str(value))
        rounded = decimal_value.quantize(
            Decimal('0.01'), 
            rounding=ROUND_HALF_UP
        )
        return float(rounded)
    
    def get_template_info(self, template_path: str) -> Dict[str, any]:
        """
        Get information about an X2Beta template.
        
        Args:
            template_path: Path to template file
            
        Returns:
            Template information
        """
        try:
            if not os.path.exists(template_path):
                return {'exists': False, 'error': 'Template file not found'}
            
            wb = load_workbook(template_path, read_only=True)
            ws = wb.active
            
            # Read headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            return {
                'exists': True,
                'sheet_name': ws.title,
                'headers': headers,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'file_size': os.path.getsize(template_path)
            }
            
        except Exception as e:
            return {
                'exists': False,
                'error': str(e)
            }
    
    def validate_template(self, template_path: str) -> Dict[str, any]:
        """
        Validate X2Beta template structure.
        
        Args:
            template_path: Path to template file
            
        Returns:
            Validation result
        """
        info = self.get_template_info(template_path)
        
        if not info.get('exists', False):
            return {
                'valid': False,
                'errors': [info.get('error', 'Unknown error')]
            }
        
        errors = []
        
        # Check for required headers
        required_headers = ['Date', 'Voucher No.', 'Party Ledger', 'Taxable Amount']
        headers = info.get('headers', [])
        
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            errors.append(f"Missing required headers: {missing_headers}")
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'headers': headers
        }
