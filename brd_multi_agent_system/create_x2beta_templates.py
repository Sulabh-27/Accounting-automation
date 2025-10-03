#!/usr/bin/env python3
"""
Create X2Beta template files for each GSTIN
"""
import sys
import os
sys.path.append('.')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_x2beta_template(gstin: str, company_name: str, state_name: str) -> str:
    """Create X2Beta template for a specific GSTIN."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Vouchers"
    
    # Company information header
    ws.merge_cells('A1:Q1')
    company_cell = ws['A1']
    company_cell.value = f"{company_name} (GSTIN: {gstin})"
    company_cell.font = Font(size=14, bold=True)
    company_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:Q2')
    state_cell = ws['A2']
    state_cell.value = f"State: {state_name}"
    state_cell.font = Font(size=12)
    state_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:Q3')
    template_cell = ws['A3']
    template_cell.value = "X2Beta Sales Import Template"
    template_cell.font = Font(size=12, bold=True)
    template_cell.alignment = Alignment(horizontal='center')
    
    # Headers row (row 4)
    headers = [
        'Date',
        'Voucher No.',
        'Voucher Type', 
        'Party Ledger',
        'Party Name',
        'Item Name',
        'Quantity',
        'Rate',
        'Taxable Amount',
        'Output CGST Ledger',
        'CGST Amount',
        'Output SGST Ledger', 
        'SGST Amount',
        'Output IGST Ledger',
        'IGST Amount',
        'Total Amount',
        'Narration'
    ]
    
    # Write headers in row 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Set column widths
    column_widths = {
        'A': 12,  # Date
        'B': 15,  # Voucher No.
        'C': 12,  # Voucher Type
        'D': 25,  # Party Ledger
        'E': 25,  # Party Name
        'F': 30,  # Item Name
        'G': 10,  # Quantity
        'H': 12,  # Rate
        'I': 15,  # Taxable Amount
        'J': 20,  # Output CGST Ledger
        'K': 12,  # CGST Amount
        'L': 20,  # Output SGST Ledger
        'M': 12,  # SGST Amount
        'N': 20,  # Output IGST Ledger
        'O': 12,  # IGST Amount
        'P': 15,  # Total Amount
        'Q': 30   # Narration
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add sample data row (row 5) for reference
    sample_data = [
        '01-08-2025',           # Date
        'AMZ-HR-08-0001',       # Voucher No.
        'Sales',                # Voucher Type
        'Amazon Haryana',       # Party Ledger
        'Amazon Haryana',       # Party Name
        'FABCON-5L',           # Item Name
        '2',                    # Quantity
        '1059.00',             # Rate
        '2118.00',             # Taxable Amount
        'Output CGST @ 18%',    # Output CGST Ledger
        '190.62',              # CGST Amount
        'Output SGST @ 18%',    # Output SGST Ledger
        '190.62',              # SGST Amount
        '',                     # Output IGST Ledger
        '0.00',                # IGST Amount
        '2499.24',             # Total Amount
        'Sales - FABCON-5L - 2025-08'  # Narration
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.alignment = Alignment(horizontal='left' if col in [3, 4, 5, 6, 10, 12, 14, 17] else 'right')
        
        # Format numbers
        if col in [7, 8, 9, 11, 13, 15, 16]:  # Numeric columns
            cell.number_format = '#,##0.00'
        elif col == 1:  # Date column
            cell.number_format = 'DD-MM-YYYY'
    
    # Add instructions in a separate sheet
    instructions_ws = wb.create_sheet("Instructions")
    
    instructions = [
        "X2Beta Sales Template Instructions",
        "",
        "1. This template is designed for importing sales data into Tally using X2Beta format",
        "2. Do not modify the header row (row 4)",
        "3. Enter data starting from row 5",
        "4. Date format: DD-MM-YYYY",
        "5. All amounts should be in INR",
        "",
        "Column Descriptions:",
        "- Date: Invoice/voucher date",
        "- Voucher No.: Unique invoice number", 
        "- Voucher Type: Always 'Sales' for sales transactions",
        "- Party Ledger: Customer ledger name in Tally",
        "- Party Name: Customer display name",
        "- Item Name: Product/service name",
        "- Quantity: Number of units sold",
        "- Rate: Price per unit",
        "- Taxable Amount: Amount before tax",
        "- Output CGST/SGST/IGST Ledger: Tax ledger names",
        "- Tax Amounts: Calculated tax values",
        "- Total Amount: Final amount including tax",
        "- Narration: Transaction description",
        "",
        f"Generated for: {company_name}",
        f"GSTIN: {gstin}",
        f"State: {state_name}"
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row, column=1, value=instruction)
        if row == 1:
            cell.font = Font(size=14, bold=True)
        elif instruction.startswith("Column Descriptions:"):
            cell.font = Font(bold=True)
        elif instruction.startswith(("Generated for:", "GSTIN:", "State:")):
            cell.font = Font(italic=True)
    
    # Auto-adjust instruction sheet column width
    instructions_ws.column_dimensions['A'].width = 80
    
    # Save template
    template_filename = f"X2Beta Sales Template - {gstin}.xlsx"
    template_path = f"ingestion_layer/templates/{template_filename}"
    
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    wb.save(template_path)
    
    return template_path


def create_all_templates():
    """Create X2Beta templates for all GSTINs."""
    
    print("🏭 Creating X2Beta Templates for All GSTINs")
    print("=" * 50)
    
    # GSTIN configurations
    gstin_configs = [
        {
            'gstin': '06ABGCS4796R1ZA',
            'company_name': 'Zaggle Haryana Private Limited',
            'state_name': 'HARYANA'
        },
        {
            'gstin': '07ABGCS4796R1Z8', 
            'company_name': 'Zaggle Delhi Private Limited',
            'state_name': 'DELHI'
        },
        {
            'gstin': '09ABGCS4796R1Z4',
            'company_name': 'Zaggle Uttar Pradesh Private Limited', 
            'state_name': 'UTTAR PRADESH'
        },
        {
            'gstin': '24ABGCS4796R1ZC',
            'company_name': 'Zaggle Gujarat Private Limited',
            'state_name': 'GUJARAT'
        },
        {
            'gstin': '29ABGCS4796R1Z2',
            'company_name': 'Zaggle Karnataka Private Limited',
            'state_name': 'KARNATAKA'
        }
    ]
    
    created_templates = []
    
    for config in gstin_configs:
        print(f"📄 Creating template for {config['gstin']} ({config['state_name']})...")
        
        try:
            template_path = create_x2beta_template(
                config['gstin'],
                config['company_name'], 
                config['state_name']
            )
            
            created_templates.append({
                'gstin': config['gstin'],
                'template_path': template_path,
                'company_name': config['company_name'],
                'state_name': config['state_name']
            })
            
            print(f"    ✅ Created: {os.path.basename(template_path)}")
            
        except Exception as e:
            print(f"    ❌ Failed to create template for {config['gstin']}: {e}")
    
    print(f"\n📊 Summary:")
    print(f"    Templates created: {len(created_templates)}")
    print(f"    Templates location: ingestion_layer/templates/")
    
    print(f"\n📋 Created Templates:")
    for template in created_templates:
        print(f"    {template['gstin']}: {os.path.basename(template['template_path'])}")
    
    return created_templates


def main():
    print("🚀 X2BETA TEMPLATE GENERATOR")
    print("Creating Excel templates for Tally X2Beta import")
    print("=" * 60)
    
    templates = create_all_templates()
    
    if len(templates) > 0:
        print(f"\n🎉 Successfully created {len(templates)} X2Beta templates!")
        print(f"📁 Templates are ready in: ingestion_layer/templates/")
        print(f"🚀 Ready for Tally export workflow")
    else:
        print(f"\n❌ Failed to create templates")
    
    return 0 if len(templates) > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
