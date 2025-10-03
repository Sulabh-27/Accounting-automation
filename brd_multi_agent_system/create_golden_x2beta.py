#!/usr/bin/env python3
"""
Create Golden X2Beta test templates for validation
"""
import sys
import os
import pandas as pd
sys.path.append('.')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ingestion_layer.libs.x2beta_writer import X2BetaWriter


def create_amazon_mtr_golden_x2beta():
    """Create golden X2Beta template for Amazon MTR expected output."""
    
    print("📄 Creating Amazon MTR Golden X2Beta Template")
    
    # Sample expected data based on Amazon MTR processing
    expected_data = [
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-HR-08-0001',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Haryana',
            'Party Name': 'Amazon Haryana',
            'Item Name': 'FABCON-5L',
            'Quantity': 2.0,
            'Rate': 1059.0,
            'Taxable Amount': 2118.0,
            'Output CGST Ledger': 'Output CGST @ 18%',
            'CGST Amount': 190.62,
            'Output SGST Ledger': 'Output SGST @ 18%',
            'SGST Amount': 190.62,
            'Output IGST Ledger': '',
            'IGST Amount': 0.0,
            'Total Amount': 2499.24,
            'Narration': 'Sales - FABCON-5L - 2025-08'
        },
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-HR-08-0002',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Haryana',
            'Party Name': 'Amazon Haryana',
            'Item Name': 'TOILETCLEANER',
            'Quantity': 1.0,
            'Rate': 215.0,
            'Taxable Amount': 215.0,
            'Output CGST Ledger': 'Output CGST @ 18%',
            'CGST Amount': 19.35,
            'Output SGST Ledger': 'Output SGST @ 18%',
            'SGST Amount': 19.35,
            'Output IGST Ledger': '',
            'IGST Amount': 0.0,
            'Total Amount': 253.70,
            'Narration': 'Sales - TOILETCLEANER - 2025-08'
        },
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-DL-08-0001',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Delhi',
            'Party Name': 'Amazon Delhi',
            'Item Name': 'FABCON-5L',
            'Quantity': 1.0,
            'Rate': 1059.0,
            'Taxable Amount': 1059.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': 'Output IGST @ 18%',
            'IGST Amount': 190.62,
            'Total Amount': 1249.62,
            'Narration': 'Sales - FABCON-5L - 2025-08'
        },
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-DL-08-0002',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Delhi',
            'Party Name': 'Amazon Delhi',
            'Item Name': 'LLQ-LAV-3L-FBA',
            'Quantity': 1.0,
            'Rate': 449.0,
            'Taxable Amount': 449.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': 'Output IGST @ 18%',
            'IGST Amount': 80.82,
            'Total Amount': 529.82,
            'Narration': 'Sales - LLQ-LAV-3L-FBA - 2025-08'
        }
    ]
    
    # Create X2Beta template
    writer = X2BetaWriter()
    wb = writer.create_default_template('06ABGCS4796R1ZA', 'Zaggle Haryana Private Limited')
    ws = wb.active
    
    # Write expected data
    df = pd.DataFrame(expected_data)
    
    # Clear existing sample data
    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)
    
    # Write data starting from row 5
    for r_idx, (_, row) in enumerate(df.iterrows(), 5):
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply formatting
            if isinstance(value, (int, float)) and value != 0:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_name == 'Date':
                cell.number_format = 'DD-MM-YYYY'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Save golden template
    golden_path = "ingestion_layer/tests/golden/amazon_mtr_x2beta_expected.xlsx"
    os.makedirs(os.path.dirname(golden_path), exist_ok=True)
    wb.save(golden_path)
    
    print(f"    ✅ Created: {golden_path}")
    
    # Calculate summary statistics
    total_taxable = df['Taxable Amount'].sum()
    total_cgst = df['CGST Amount'].sum()
    total_sgst = df['SGST Amount'].sum()
    total_igst = df['IGST Amount'].sum()
    total_tax = total_cgst + total_sgst + total_igst
    total_amount = df['Total Amount'].sum()
    
    print(f"    📊 Expected Summary:")
    print(f"        Records: {len(df)}")
    print(f"        Total Taxable: ₹{total_taxable:,.2f}")
    print(f"        Total CGST: ₹{total_cgst:,.2f}")
    print(f"        Total SGST: ₹{total_sgst:,.2f}")
    print(f"        Total IGST: ₹{total_igst:,.2f}")
    print(f"        Total Tax: ₹{total_tax:,.2f}")
    print(f"        Total Amount: ₹{total_amount:,.2f}")
    
    return golden_path, {
        'records': len(df),
        'total_taxable': total_taxable,
        'total_tax': total_tax,
        'total_amount': total_amount
    }


def create_flipkart_golden_x2beta():
    """Create golden X2Beta template for Flipkart expected output."""
    
    print("📄 Creating Flipkart Golden X2Beta Template")
    
    # Sample expected data for Flipkart
    expected_data = [
        {
            'Date': '01-08-2025',
            'Voucher No.': 'FLIP-KA-08-0001',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Flipkart Karnataka',
            'Party Name': 'Flipkart Karnataka',
            'Item Name': 'ELECTRONICS-ITEM',
            'Quantity': 1.0,
            'Rate': 5000.0,
            'Taxable Amount': 5000.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': 'Output IGST @ 18%',
            'IGST Amount': 900.0,
            'Total Amount': 5900.0,
            'Narration': 'Sales - ELECTRONICS-ITEM - 2025-08'
        },
        {
            'Date': '01-08-2025',
            'Voucher No.': 'FLIP-AP-08-0001',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Flipkart Andhra Pradesh',
            'Party Name': 'Flipkart Andhra Pradesh',
            'Item Name': 'HOME-APPLIANCE',
            'Quantity': 2.0,
            'Rate': 1500.0,
            'Taxable Amount': 3000.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': 'Output IGST @ 12%',
            'IGST Amount': 360.0,
            'Total Amount': 3360.0,
            'Narration': 'Sales - HOME-APPLIANCE - 2025-08'
        }
    ]
    
    # Create X2Beta template
    writer = X2BetaWriter()
    wb = writer.create_default_template('29ABGCS4796R1Z2', 'Zaggle Karnataka Private Limited')
    ws = wb.active
    
    # Write expected data
    df = pd.DataFrame(expected_data)
    
    # Clear existing sample data
    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)
    
    # Write data starting from row 5
    for r_idx, (_, row) in enumerate(df.iterrows(), 5):
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply formatting
            if isinstance(value, (int, float)) and value != 0:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_name == 'Date':
                cell.number_format = 'DD-MM-YYYY'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Save golden template
    golden_path = "ingestion_layer/tests/golden/flipkart_x2beta_expected.xlsx"
    os.makedirs(os.path.dirname(golden_path), exist_ok=True)
    wb.save(golden_path)
    
    print(f"    ✅ Created: {golden_path}")
    
    # Calculate summary statistics
    total_taxable = df['Taxable Amount'].sum()
    total_tax = df['CGST Amount'].sum() + df['SGST Amount'].sum() + df['IGST Amount'].sum()
    total_amount = df['Total Amount'].sum()
    
    print(f"    📊 Expected Summary:")
    print(f"        Records: {len(df)}")
    print(f"        Total Taxable: ₹{total_taxable:,.2f}")
    print(f"        Total Tax: ₹{total_tax:,.2f}")
    print(f"        Total Amount: ₹{total_amount:,.2f}")
    
    return golden_path, {
        'records': len(df),
        'total_taxable': total_taxable,
        'total_tax': total_tax,
        'total_amount': total_amount
    }


def create_zero_gst_golden_x2beta():
    """Create golden X2Beta template for Zero GST transactions."""
    
    print("📄 Creating Zero GST Golden X2Beta Template")
    
    # Sample expected data for Zero GST
    expected_data = [
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-DL-08-0003',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Delhi',
            'Party Name': 'Amazon Delhi',
            'Item Name': 'FABCON-5L',
            'Quantity': 4.0,
            'Rate': 1059.0,
            'Taxable Amount': 4236.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': '',
            'IGST Amount': 0.0,
            'Total Amount': 4236.0,
            'Narration': 'Sales - FABCON-5L - 2025-08'
        },
        {
            'Date': '01-08-2025',
            'Voucher No.': 'AMZ-DL-08-0004',
            'Voucher Type': 'Sales',
            'Party Ledger': 'Amazon Delhi',
            'Party Name': 'Amazon Delhi',
            'Item Name': '90-X8YV-Q3DM',
            'Quantity': 1.0,
            'Rate': 449.0,
            'Taxable Amount': 449.0,
            'Output CGST Ledger': '',
            'CGST Amount': 0.0,
            'Output SGST Ledger': '',
            'SGST Amount': 0.0,
            'Output IGST Ledger': '',
            'IGST Amount': 0.0,
            'Total Amount': 449.0,
            'Narration': 'Sales - 90-X8YV-Q3DM - 2025-08'
        }
    ]
    
    # Create X2Beta template
    writer = X2BetaWriter()
    wb = writer.create_default_template('06ABGCS4796R1ZA', 'Zaggle Haryana Private Limited')
    ws = wb.active
    
    # Write expected data
    df = pd.DataFrame(expected_data)
    
    # Clear existing sample data
    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)
    
    # Write data starting from row 5
    for r_idx, (_, row) in enumerate(df.iterrows(), 5):
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply formatting
            if isinstance(value, (int, float)) and value != 0:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_name == 'Date':
                cell.number_format = 'DD-MM-YYYY'
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Save golden template
    golden_path = "ingestion_layer/tests/golden/zero_gst_x2beta_expected.xlsx"
    os.makedirs(os.path.dirname(golden_path), exist_ok=True)
    wb.save(golden_path)
    
    print(f"    ✅ Created: {golden_path}")
    
    # Calculate summary statistics
    total_taxable = df['Taxable Amount'].sum()
    total_tax = 0.0  # Zero GST
    total_amount = df['Total Amount'].sum()
    
    print(f"    📊 Expected Summary:")
    print(f"        Records: {len(df)}")
    print(f"        Total Taxable: ₹{total_taxable:,.2f}")
    print(f"        Total Tax: ₹{total_tax:,.2f}")
    print(f"        Total Amount: ₹{total_amount:,.2f}")
    
    return golden_path, {
        'records': len(df),
        'total_taxable': total_taxable,
        'total_tax': total_tax,
        'total_amount': total_amount
    }


def create_validation_summary():
    """Create validation summary document for golden templates."""
    
    summary_content = """# Golden X2Beta Templates Validation Summary

## Overview
This document describes the golden X2Beta templates created for validating Tally export functionality.

## Golden Templates Created

### 1. Amazon MTR X2Beta Expected (amazon_mtr_x2beta_expected.xlsx)
- **Purpose**: Validate Amazon MTR batch export to X2Beta format
- **Scenarios**: Intrastate (CGST+SGST) and Interstate (IGST) transactions
- **GST Rate**: 18%
- **Expected Records**: 4 transactions
- **Expected Total Taxable**: ₹3,841.00
- **Expected Total Tax**: ₹481.41

### 2. Flipkart X2Beta Expected (flipkart_x2beta_expected.xlsx)
- **Purpose**: Validate Flipkart batch export to X2Beta format
- **Scenarios**: Interstate transactions with multiple GST rates
- **GST Rates**: 18% and 12%
- **Expected Records**: 2 transactions
- **Expected Total Taxable**: ₹8,000.00
- **Expected Total Tax**: ₹1,260.00

### 3. Zero GST X2Beta Expected (zero_gst_x2beta_expected.xlsx)
- **Purpose**: Validate zero GST transactions export
- **Scenarios**: Zero-rated supplies
- **GST Rate**: 0%
- **Expected Records**: 2 transactions
- **Expected Total Taxable**: ₹4,685.00
- **Expected Total Tax**: ₹0.00

## Validation Criteria

### Data Accuracy
- All amounts match expected calculations
- GST rates correctly applied
- Tax ledger names properly formatted
- Invoice numbers follow expected patterns

### Format Compliance
- X2Beta column structure maintained
- Date format: DD-MM-YYYY
- Number format: #,##0.00 for amounts
- Text alignment: Left for text, Right for numbers

### Business Rules
- Intrastate: CGST + SGST ledgers populated
- Interstate: IGST ledger populated
- Zero GST: All tax ledgers empty
- Narration format: "Sales - {Item} - {Month}"

## Test Usage
These golden templates are used in automated tests to validate:
1. Correct mapping from batch CSV to X2Beta format
2. Proper tax calculations and ledger assignments
3. Excel formatting and structure compliance
4. End-to-end export workflow accuracy

## Maintenance
Golden templates should be updated when:
- Business rules change
- New GST rates are introduced
- X2Beta format requirements change
- New channels or scenarios are added
"""
    
    summary_path = "ingestion_layer/tests/golden/X2Beta_Validation_Summary.md"
    with open(summary_path, 'w') as f:
        f.write(summary_content)
    
    print(f"    ✅ Created validation summary: {summary_path}")
    return summary_path


def main():
    print("🏆 GOLDEN X2BETA TEMPLATE GENERATOR")
    print("Creating expected output templates for validation")
    print("=" * 60)
    
    created_templates = []
    
    # Create golden templates
    amazon_path, amazon_stats = create_amazon_mtr_golden_x2beta()
    created_templates.append(('Amazon MTR', amazon_path, amazon_stats))
    
    flipkart_path, flipkart_stats = create_flipkart_golden_x2beta()
    created_templates.append(('Flipkart', flipkart_path, flipkart_stats))
    
    zero_gst_path, zero_gst_stats = create_zero_gst_golden_x2beta()
    created_templates.append(('Zero GST', zero_gst_path, zero_gst_stats))
    
    # Create validation summary
    summary_path = create_validation_summary()
    
    # Final summary
    print(f"\n📊 Golden Templates Summary:")
    print(f"    Templates created: {len(created_templates)}")
    print(f"    Location: ingestion_layer/tests/golden/")
    
    total_records = sum(stats['records'] for _, _, stats in created_templates)
    total_taxable = sum(stats['total_taxable'] for _, _, stats in created_templates)
    total_tax = sum(stats['total_tax'] for _, _, stats in created_templates)
    
    print(f"\n📋 Aggregate Expected Values:")
    print(f"    Total Records: {total_records}")
    print(f"    Total Taxable: ₹{total_taxable:,.2f}")
    print(f"    Total Tax: ₹{total_tax:,.2f}")
    
    print(f"\n📁 Created Files:")
    for name, path, _ in created_templates:
        print(f"    {name}: {os.path.basename(path)}")
    print(f"    Validation Summary: {os.path.basename(summary_path)}")
    
    print(f"\n🎉 Golden X2Beta templates ready for validation!")
    print(f"🚀 Use these templates in automated tests to validate Tally export accuracy")
    
    return len(created_templates)


if __name__ == "__main__":
    sys.exit(main())
