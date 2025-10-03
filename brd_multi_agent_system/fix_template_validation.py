#!/usr/bin/env python3
"""
Fix template validation and test Part-5
"""
import sys
import os
import pandas as pd
sys.path.append('.')

from openpyxl import load_workbook


def check_template_structure():
    """Check the actual structure of X2Beta template."""
    
    print("🔍 CHECKING X2BETA TEMPLATE STRUCTURE")
    print("=" * 50)
    
    template_path = "ingestion_layer/templates/X2Beta Sales Template - 06ABGCS4796R1ZA.xlsx"
    
    if not os.path.exists(template_path):
        print(f"❌ Template not found: {template_path}")
        return False
    
    try:
        # Load workbook
        wb = load_workbook(template_path, read_only=True)
        ws = wb.active
        
        print(f"✅ Template loaded successfully")
        print(f"    Sheet name: {ws.title}")
        print(f"    Max row: {ws.max_row}")
        print(f"    Max column: {ws.max_column}")
        
        # Check first few rows
        print(f"\n📋 Template Structure:")
        for row_num in range(1, min(8, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, min(18, ws.max_column + 1)):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    row_data.append(str(cell_value)[:20])
                else:
                    row_data.append("")
            
            if any(row_data):  # Only show non-empty rows
                print(f"    Row {row_num}: {row_data[:5]}...")  # Show first 5 columns
        
        # Find header row
        print(f"\n🔍 Looking for header row:")
        for row_num in range(1, min(10, ws.max_row + 1)):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    row_values.append(str(cell_value))
            
            if 'Date' in row_values and 'Voucher No.' in row_values:
                print(f"    ✅ Found headers in row {row_num}: {row_values[:8]}...")
                return row_num
        
        print(f"    ❌ Could not find header row with Date and Voucher No.")
        return False
        
    except Exception as e:
        print(f"❌ Error checking template: {e}")
        return False


def test_x2beta_export_simple():
    """Simple test of X2Beta export functionality."""
    
    print(f"\n🏭 TESTING X2BETA EXPORT (SIMPLE)")
    print("=" * 50)
    
    # Check batch files
    batch_dir = "ingestion_layer/data/batches"
    batch_files = [f for f in os.listdir(batch_dir) if f.endswith('_batch.csv')]
    
    if not batch_files:
        print("❌ No batch files found")
        return False
    
    # Load a batch file
    batch_file = os.path.join(batch_dir, batch_files[0])
    print(f"📄 Using batch file: {os.path.basename(batch_file)}")
    
    try:
        df = pd.read_csv(batch_file)
        print(f"✅ Loaded {len(df)} records")
        print(f"    Columns: {list(df.columns)}")
        
        if len(df) > 0:
            sample_record = df.iloc[0]
            print(f"    Sample record:")
            print(f"      GST Rate: {sample_record.get('gst_rate', 'N/A')}")
            print(f"      Ledger: {sample_record.get('ledger_name', 'N/A')}")
            print(f"      FG: {sample_record.get('fg', 'N/A')}")
            print(f"      Taxable: ₹{sample_record.get('total_taxable', 0):,.2f}")
        
        # Test X2Beta mapping
        from ingestion_layer.libs.x2beta_writer import X2BetaWriter
        
        writer = X2BetaWriter()
        x2beta_df = writer.map_batch_to_x2beta(df, '06ABGCS4796R1ZA', '2025-08')
        
        print(f"✅ X2Beta mapping successful: {len(x2beta_df)} records")
        
        if len(x2beta_df) > 0:
            sample_x2beta = x2beta_df.iloc[0]
            print(f"    X2Beta sample:")
            print(f"      Date: {sample_x2beta.get('Date', 'N/A')}")
            print(f"      Voucher No.: {sample_x2beta.get('Voucher No.', 'N/A')}")
            print(f"      Party Ledger: {sample_x2beta.get('Party Ledger', 'N/A')}")
            print(f"      Taxable Amount: ₹{sample_x2beta.get('Taxable Amount', 0):,.2f}")
            
            # Check tax amounts
            cgst = sample_x2beta.get('CGST Amount', 0)
            sgst = sample_x2beta.get('SGST Amount', 0)
            igst = sample_x2beta.get('IGST Amount', 0)
            
            if cgst > 0 or sgst > 0:
                print(f"      Tax Type: Intrastate (CGST: ₹{cgst:.2f}, SGST: ₹{sgst:.2f})")
            elif igst > 0:
                print(f"      Tax Type: Interstate (IGST: ₹{igst:.2f})")
            else:
                print(f"      Tax Type: Zero GST")
        
        # Test writing to Excel (simplified)
        print(f"\n📄 Testing Excel export:")
        
        # Create a simple output file
        output_dir = "ingestion_layer/exports"
        os.makedirs(output_dir, exist_ok=True)
        
        output_file = os.path.join(output_dir, "test_x2beta_export.xlsx")
        
        # Save as Excel
        x2beta_df.to_excel(output_file, index=False, sheet_name="Sales Vouchers")
        
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"✅ Excel file created: {os.path.basename(output_file)} ({file_size:,} bytes)")
            
            # Verify by reading back
            verify_df = pd.read_excel(output_file)
            print(f"✅ Verification: {len(verify_df)} records read back")
            
            return True
        else:
            print(f"❌ Excel file not created")
            return False
        
    except Exception as e:
        print(f"❌ Error in X2Beta export test: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    print("🚀 PART-5 TEMPLATE VALIDATION & EXPORT TEST")
    print("=" * 60)
    
    # Step 1: Check template structure
    header_row = check_template_structure()
    
    # Step 2: Test X2Beta export
    export_success = test_x2beta_export_simple()
    
    # Summary
    print(f"\n" + "=" * 60)
    print("📋 TEST SUMMARY")
    print("=" * 60)
    
    if header_row and export_success:
        print("✅ Part-5 components are working!")
        print("✅ X2Beta templates are properly structured")
        print("✅ Batch to X2Beta conversion successful")
        print("✅ Excel export working")
        print(f"\n🚀 Ready for full pipeline test!")
    else:
        print("❌ Some issues found:")
        if not header_row:
            print("  - Template structure needs verification")
        if not export_success:
            print("  - X2Beta export needs debugging")
    
    return 0 if (header_row and export_success) else 1


if __name__ == "__main__":
    sys.exit(main())
