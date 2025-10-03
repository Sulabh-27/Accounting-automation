#!/usr/bin/env python3
"""
Create X2Beta Expense Templates for all GSTINs
Templates for expense/purchase vouchers in Tally-compatible format
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_expense_template(gstin: str, company_name: str, state_name: str, output_path: str):
    """Create X2Beta expense template for a specific GSTIN."""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Vouchers"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Company header information
    ws['A1'] = f"{company_name}"
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A2'] = f"GSTIN: {gstin}"
    ws['A2'].font = Font(name='Arial', size=11, bold=True)
    ws['A3'] = f"State: {state_name}"
    ws['A3'].font = Font(name='Arial', size=11)
    ws['A4'] = "X2Beta Expense Import Template"
    ws['A4'].font = Font(name='Arial', size=12, bold=True, color='FF0000')
    
    # Column headers (starting from row 6)
    headers = [
        'Date',
        'Voucher No.',
        'Voucher Type', 
        'Party Ledger',
        'Item Name',
        'Quantity',
        'Rate',
        'Taxable Amount',
        'CGST Amount',
        'SGST Amount', 
        'IGST Amount',
        'Total Amount',
        'Narration'
    ]
    
    # Set headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Sample data rows (rows 7-10)
    sample_data = [
        # Expense entry (debit)
        ['01-08-2025', 'EXP060825001', 'Purchase', 'Amazon Commission', 'Commission - AMZ-FEE-001', 1, 1000.00, 1000.00, 0, 0, 0, 1000.00, 'Amazon marketplace commission'],
        # Input IGST entry (debit) 
        ['01-08-2025', 'EXP060825001', 'Purchase', 'Input IGST @ 18%', 'Input IGST - AMZ-FEE-001', 1, 180.00, 0, 0, 0, 180.00, 180.00, 'Input IGST on commission'],
        # Vendor payable entry (credit)
        ['01-08-2025', 'EXP060825001', 'Purchase', 'Amazon Payable', 'Payable - AMZ-FEE-001', 1, -1180.00, -1180.00, 0, 0, 0, -1180.00, 'Amount payable to Amazon'],
        # Another expense example
        ['02-08-2025', 'EXP060825002', 'Purchase', 'Amazon Shipping Fee', 'Shipping - AMZ-FEE-002', 1, 500.00, 500.00, 0, 0, 0, 500.00, 'Amazon shipping charges']
    ]
    
    # Add sample data
    for row_idx, row_data in enumerate(sample_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Format numbers
            if col_idx in [6, 7, 8, 9, 10, 11, 12]:  # Amount columns
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12,  # Date
        'B': 15,  # Voucher No
        'C': 12,  # Voucher Type
        'D': 25,  # Party Ledger
        'E': 30,  # Item Name
        'F': 10,  # Quantity
        'G': 12,  # Rate
        'H': 15,  # Taxable Amount
        'I': 12,  # CGST Amount
        'J': 12,  # SGST Amount
        'K': 12,  # IGST Amount
        'L': 15,  # Total Amount
        'M': 35   # Narration
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    
    instructions = [
        "X2Beta Expense Template Instructions",
        "",
        "This template is used to import expense/purchase vouchers into Tally ERP.",
        "",
        "Column Descriptions:",
        "- Date: Transaction date (DD-MM-YYYY format)",
        "- Voucher No.: Unique voucher number for the transaction",
        "- Voucher Type: Usually 'Purchase' for expense entries",
        "- Party Ledger: Name of the ledger account (expense account or GST ledger)",
        "- Item Name: Description of the expense item",
        "- Quantity: Usually 1 for expenses",
        "- Rate: Amount per unit",
        "- Taxable Amount: Amount before GST (for expense ledgers)",
        "- CGST Amount: Central GST amount",
        "- SGST Amount: State GST amount", 
        "- IGST Amount: Integrated GST amount",
        "- Total Amount: Final amount (positive for debits, negative for credits)",
        "- Narration: Description/remarks for the entry",
        "",
        "Important Notes:",
        "- Each expense transaction requires multiple entries:",
        "  1. Debit to expense ledger (taxable amount)",
        "  2. Debit to input GST ledgers (GST amounts)",
        "  3. Credit to vendor/payable ledger (total amount as negative)",
        "",
        "- For intrastate transactions: Use CGST + SGST",
        "- For interstate transactions: Use IGST",
        "",
        "- Ensure the sum of all amounts in a voucher equals zero",
        "- Use consistent voucher numbers for related entries",
        "",
        "GST Ledger Naming Convention:",
        "- Input CGST @ 9% (for 18% GST rate)",
        "- Input SGST @ 9% (for 18% GST rate)", 
        "- Input IGST @ 18% (for 18% GST rate)",
        "",
        f"This template is configured for {company_name} (GSTIN: {gstin})",
        f"State: {state_name}"
    ]
    
    for row, instruction in enumerate(instructions, 1):
        instructions_ws[f'A{row}'] = instruction
        if row == 1:
            instructions_ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        elif instruction.endswith(':'):
            instructions_ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
    
    # Adjust column width for instructions
    instructions_ws.column_dimensions['A'].width = 80
    
    # Save the workbook
    wb.save(output_path)
    print(f"Created expense template: {output_path}")


def create_all_expense_templates():
    """Create X2Beta expense templates for all GSTINs."""
    
    # Company configurations
    companies = [
        {
            'gstin': '06ABGCS4796R1ZA',
            'company_name': 'Zaggle Haryana Private Limited',
            'state_name': 'HARYANA'
        },
        {
            'gstin': '07ABGCS4796R1Z8', 
            'company_name': 'Zaggle Delhi Private Limited',
            'state_name': 'DELHI'
        },
        {
            'gstin': '09ABGCS4796R1Z4',
            'company_name': 'Zaggle Uttar Pradesh Private Limited', 
            'state_name': 'UTTAR PRADESH'
        },
        {
            'gstin': '24ABGCS4796R1ZC',
            'company_name': 'Zaggle Gujarat Private Limited',
            'state_name': 'GUJARAT'
        },
        {
            'gstin': '29ABGCS4796R1Z2',
            'company_name': 'Zaggle Karnataka Private Limited',
            'state_name': 'KARNATAKA'
        }
    ]
    
    # Create templates directory if it doesn't exist
    templates_dir = "ingestion_layer/templates"
    os.makedirs(templates_dir, exist_ok=True)
    
    print("Creating X2Beta Expense Templates...")
    print("=" * 50)
    
    for company in companies:
        template_name = f"X2Beta Expense Template - {company['gstin']}.xlsx"
        template_path = os.path.join(templates_dir, template_name)
        
        create_expense_template(
            company['gstin'],
            company['company_name'], 
            company['state_name'],
            template_path
        )
    
    print("\n✅ All expense templates created successfully!")
    print(f"📁 Templates saved in: {templates_dir}")
    
    # List created templates
    print("\n📋 Created Templates:")
    for company in companies:
        template_name = f"X2Beta Expense Template - {company['gstin']}.xlsx"
        print(f"  - {template_name} ({company['company_name']})")


if __name__ == "__main__":
    create_all_expense_templates()
